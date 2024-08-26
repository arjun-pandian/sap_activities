import sys
import os
current_directory = os.getcwd()
functions_directory = os.path.abspath(os.path.join(current_directory, 'Files', 'Common'))
sys.path.append(functions_directory)

from sap_functions import sap_login, create_session, take_and_save_screenshot, today, yesterday
import time
import pyautogui
import re

def run_pk9_script():
    sap_login("PK9", "215")

    session = create_session()
    pyautogui.hotkey('win','up')

    session.findById("wnd[0]").maximize()

    #'''
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nsm12"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/usr/txtSEQG3-GUNAME").text = "*"
    session.findById("wnd[0]").sendVKey(8)
    take_and_save_screenshot("pk9", "1_sm12")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").setCurrentCell (-1,"GDSPTIME")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").selectColumn ("GDSPTIME")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").contextMenu()
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").selectContextMenuItem ("&FIND")
    session.findById("wnd[1]/usr/chkGS_SEARCH-EXACT_WORD").selected = True
    session.findById("wnd[1]/usr/chkGS_SEARCH-SHOW_HITS").selected = True
    session.findById("wnd[1]/usr/txtGS_SEARCH-VALUE").text = "*.*.*"
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    lock_details_1 = session.findById("wnd[1]/usr/txtGS_SEARCH-SEARCH_INFO").text

    if "No Hit Found" in lock_details_1 :
        sm12_previous_locks = 0
    else :
        numbers = re.findall(r'\d+', lock_details_1)
        sm12_previous_locks = numbers[-1] 

    #print("Previous lock :", sm12_previous_locks)
    session.findById("wnd[1]").close()

    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").setCurrentCell (-1,"GDSPTIME")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").selectColumn ("GDSPTIME")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").contextMenu()
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").selectContextMenuItem ("&FIND")
    session.findById("wnd[1]/usr/chkGS_SEARCH-EXACT_WORD").selected = True
    session.findById("wnd[1]/usr/chkGS_SEARCH-SHOW_HITS").selected = True
    session.findById("wnd[1]/usr/txtGS_SEARCH-VALUE").text = "*:*:*"
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    lock_details_2 = session.findById("wnd[1]/usr/txtGS_SEARCH-SEARCH_INFO").text
    
    if "No Hit Found" in lock_details_2 :
        sm12_today_locks = 0
    else:
        numbers = re.findall(r'\d+', lock_details_2)
        sm12_today_locks = numbers[-1] 

    #print("Today's lock :",sm12_today_locks)
    session.findById("wnd[1]").close()

    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").setCurrentCell (-1,"GDSPTIME")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").selectColumn ("GDSPTIME")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").contextMenu()
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").selectContextMenuItem ("&FIND")
    session.findById("wnd[1]/usr/chkGS_SEARCH-EXACT_WORD").selected = True
    session.findById("wnd[1]/usr/chkGS_SEARCH-SHOW_HITS").selected = True
    session.findById("wnd[1]/usr/txtGS_SEARCH-VALUE").text = "*"
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    lock_details_3 = session.findById("wnd[1]/usr/txtGS_SEARCH-SEARCH_INFO").text
    if "No Hit Found" in lock_details_3 :
        sm12_total_locks = 0
    else:
        numbers = re.findall(r'\d+', lock_details_3)
        sm12_total_locks = numbers[-1] 
    
    #print("Total lock :",sm12_total_locks)
    session.findById("wnd[1]").close()



    session.findById("wnd[0]/tbar[0]/okcd").text = "/nsm13"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/usr/radSEL_STOPPED").select()
    session.findById("wnd[0]").sendVKey(8)
    take_and_save_screenshot("pk9", "2_sm13")

    session.findById("wnd[0]/tbar[0]/okcd").text = "/nsm21"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]").sendVKey(8)
    take_and_save_screenshot("pk9", "3_sm21")
    
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nst22"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/usr/ctxtS_DATUM-LOW").text = yesterday
    session.findById("wnd[0]/usr/ctxtS_DATUM-HIGH").text = today
    session.findById("wnd[0]/usr/ctxtS_UZEIT-LOW").text = ""
    session.findById("wnd[0]/usr/ctxtS_UZEIT-HIGH").text = ""
    session.findById("wnd[0]/usr/txtS_UNAME-LOW").text = "*"
    session.findById("wnd[0]/usr/txtTOD_NUM").setFocus()
    runtime_error_details = session.findById("wnd[0]/usr/txtTOD_NUM").Text
    st22_today_runtime_errors = re.search(r'\d+', runtime_error_details).group()
    #print("Today Runtime Errors :",st22_today_runtime_errors)
    session.findById("wnd[0]/usr/btnSTARTSEL").press()
    session.findById("wnd[0]/usr/cntlRSSHOWRABAX_ALV_100/shellcont/shell").setCurrentCell (-1,"ERRORID")
    session.findById("wnd[0]/usr/cntlRSSHOWRABAX_ALV_100/shellcont/shell").selectColumn ("ERRORID")
    session.findById("wnd[0]/usr/cntlRSSHOWRABAX_ALV_100/shellcont/shell").contextMenu()
    session.findById("wnd[0]/usr/cntlRSSHOWRABAX_ALV_100/shellcont/shell").pressToolbarButton ("&SORT_DSC")
    take_and_save_screenshot("pk9", "4_st22")
    #'''

    session.findById("wnd[0]/tbar[0]/okcd").text = "/nsm37"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/usr/chkBTCH2170-PRELIM").selected = False
    session.findById("wnd[0]/usr/chkBTCH2170-SCHEDUL").selected = False
    session.findById("wnd[0]/usr/chkBTCH2170-READY").selected = True
    session.findById("wnd[0]/usr/chkBTCH2170-RUNNING").selected = True
    session.findById("wnd[0]/usr/chkBTCH2170-FINISHED").selected = False
    session.findById("wnd[0]/usr/chkBTCH2170-ABORTED").selected = False
    session.findById("wnd[0]/usr/txtBTCH2170-USERNAME").text = "*"
    session.findById("wnd[0]/usr/ctxtBTCH2170-FROM_DATE").text = ""
    session.findById("wnd[0]/usr/ctxtBTCH2170-TO_DATE").text = ""
    session.findById("wnd[0]").sendVKey(8)
    try:
        session.findById("wnd[0]/tbar[0]/btn[83]").press()
        session.findById("wnd[0]/usr/lbl[102,10]").setFocus()
        session.findById("wnd[0]").sendVKey (2)
        session.findById("wnd[0]").sendVKey (41)
        take_and_save_screenshot("pk9", "5_sm37_1")

        total_job_time = 0
        for x in range(12, 30):
            try:
                element = session.findById(f"wnd[0]/usr/lbl[4,{x}]")
                text = element.Text

                if "Summary" in text:
                    job_time_details = session.findById(f"wnd[0]/usr/lbl[102,{x}]").Text
                    sm37_total_job_time = job_time_details.replace(' ', '').replace('.', '')
                    #print("Total job time :", sm37_total_job_time)
                    break 

            except Exception as e:
                #print(f"Error occurred while searching for index {x}")
                continue 

        else:
            print("Summary not found within the specified range")
    except:
        take_and_save_screenshot("pk9", "5_sm37_1")

    session.findById("wnd[0]").sendVKey(3)
    session.findById("wnd[0]/usr/chkBTCH2170-RUNNING").selected = False
    session.findById("wnd[0]/usr/chkBTCH2170-ABORTED").selected = True
    session.findById("wnd[0]/usr/chkBTCH2170-READY").selected = False
    session.findById("wnd[0]/usr/ctxtBTCH2170-FROM_DATE").text = yesterday
    session.findById("wnd[0]/usr/ctxtBTCH2170-TO_DATE").text = today
    session.findById("wnd[0]/usr/ctxtBTCH2170-FROM_TIME").text = ""
    session.findById("wnd[0]/usr/ctxtBTCH2170-TO_TIME").text = ""
    session.findById("wnd[0]").sendVKey(8)
    try:
        session.findById("wnd[0]/tbar[0]/btn[83]").press()
        session.findById("wnd[0]/usr/lbl[102,10]").setFocus()
        session.findById("wnd[0]").sendVKey (2)
        session.findById("wnd[0]").sendVKey (41)    
        take_and_save_screenshot("pk9", "6_sm37_2")
    except:
        take_and_save_screenshot("pk9", "6_sm37_2")

    #'''
    session.findById("wnd[0]/tbar[0]/okcd").text = "/nsm51"
    session.findById("wnd[0]").sendVKey (0)
    take_and_save_screenshot("pk9", "7_sm51")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").setCurrentCell (-1,"STATUS")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").selectColumn ("STATUS")
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").contextMenu()
    session.findById("wnd[0]/usr/cntlGRID1/shellcont/shell/shellcont[0]/shell").selectContextMenuItem ("&FIND")
    session.findById("wnd[1]/usr/chkGS_SEARCH-EXACT_WORD").selected = True
    session.findById("wnd[1]/usr/chkGS_SEARCH-SHOW_HITS").selected = True
    session.findById("wnd[1]/usr/txtGS_SEARCH-VALUE").text = "*"
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    status_details = session.findById("wnd[1]/usr/txtGS_SEARCH-SEARCH_INFO").text
    try : 
        numbers = re.findall(r'\d+', status_details)
        total = numbers[-1] 
    except :
        total = 0
    #print("Total :", total)

    session.findById("wnd[1]/usr/txtGS_SEARCH-VALUE").text = "Active"
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    status_details = session.findById("wnd[1]/usr/txtGS_SEARCH-SEARCH_INFO").text
    try : 
        numbers = re.findall(r'\d+', status_details)
        total_active = numbers[-1] 
    except :
        total_active = 0
    #print("Total Active :", total_active)

    sm51_status = ""

    if total == total_active :
        sm51_status = "All are active"
    else :
        sm51_status = "Not active"

    session.findById("wnd[1]").close()


    session.findById("wnd[0]/tbar[0]/okcd").text = "/nsm66"
    session.findById("wnd[0]").sendVKey (0)
    take_and_save_screenshot("pk9", "8_sm66")


    session.findById("wnd[0]/tbar[0]/okcd").text = "/nsmq1"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]").sendVKey(8)
    take_and_save_screenshot("pk9", "9_smq1")


    session.findById("wnd[0]/tbar[0]/okcd").text = "/nstms"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]").sendVKey(5)
    session.findById("wnd[0]/usr/lbl[2,7]").setFocus()
    hovered_cell = session.findById("wnd[0]/usr/lbl[2,7]").Text
    take_and_save_screenshot("pk9", "10_stms_1")

    if hovered_cell == "PK9":
        #print("PK9")
        session.findById("wnd[0]/tbar[1]/btn[18]").press()
        session.findById("wnd[1]/usr/btnSPOP-OPTION2").press()
    else:
        print("Not a PK9")
        for x in range(6, 30):
            try:
                element = session.findById(f"wnd[0]/usr/lbl[2,{x}]")
                text = element.Text

                if "PK9" in text:
                    take_and_save_screenshot("pk9", "10_stms_1")
                    session.findById("wnd[0]/tbar[1]/btn[18]").press()
                    session.findById("wnd[1]/usr/btnSPOP-OPTION2").press()
                    print("PK9 found")
                    break 
            except Exception as e:
                print(f"Error occurred while searching for index {x}")
            continue 

    take_and_save_screenshot("pk9", "11_stms_2")


    session.findById("wnd[0]/tbar[0]/okcd").text = "/nsm58"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/usr/txtBENUTZER-LOW").text = "*"
    session.findById("wnd[0]").sendVKey(8)
    take_and_save_screenshot("pk9", "12_sm58")


    session.findById("wnd[0]/tbar[0]/okcd").text = "/ndb02"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/shellcont[0]").dockerPixelSize = 38
    session.findById("wnd[0]/usr/txtDB02N_DATA-DB_TOT_PER_USED").setFocus()
    db02_percentage = session.findById("wnd[0]/usr/txtDB02N_DATA-DB_TOT_PER_USED").Text
    #print("db02 : ",db02_percentage)
    take_and_save_screenshot("pk9", "13_db02")


    session.findById("wnd[0]/tbar[0]/okcd").text = "/ndb12"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/shellcont").dockerPixelSize = 38
    session.findById("wnd[0]/usr/btnTOTAL_BUTTON").press()
    take_and_save_screenshot("pk9", "14_db12")


    session.findById("wnd[0]/tbar[0]/okcd").text = "/nal08"
    session.findById("wnd[0]").sendVKey (0)
    take_and_save_screenshot("pk9", "15_al08")


    session.findById("wnd[0]/tbar[0]/okcd").text = "/nal11"
    session.findById("wnd[0]").sendVKey (0)
    take_and_save_screenshot("pk9", "16_al11")


    session.findById("wnd[0]/tbar[0]/okcd").text = "/ndb02"
    session.findById("wnd[0]").sendVKey (0)
    session.findById("wnd[0]/shellcont[1]/shell/shellcont[1]/shell").hierarchyHeaderWidth = 398
    session.findById("wnd[0]/shellcont[1]/shell/shellcont[1]/shell").expandNode ("       1008")
    session.findById("wnd[0]/shellcont[1]/shell/shellcont[1]/shell").topNode = "       1009-"
    session.findById("wnd[0]/shellcont[1]/shell/shellcont[1]/shell").selectItem ("         47","Task")
    session.findById("wnd[0]/shellcont[1]/shell/shellcont[1]/shell").ensureVisibleHorizontalItem ("         47","Task")
    session.findById("wnd[0]/shellcont[1]/shell/shellcont[1]/shell").doubleClickItem ("         47","Task")
    session.findById("wnd[0]/shellcont[0]").dockerPixelSize = 38
    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").setCurrentCell (-1,"TOT_PER_USED")
    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").selectColumn ("TOT_PER_USED")
    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").contextMenu()
    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").selectContextMenuItem ("&SORT_DSC")
    take_and_save_screenshot("pk9", "17_tablespace") 

    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").currentCellColumn = "TABLESPACE"
    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").selectColumn ("TABLESPACE")
    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").contextMenu()
    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").selectContextMenuItem ("&FILTER")
    session.findById("wnd[1]/tbar[0]/btn[2]").press()
    session.findById("wnd[2]/usr/cntlOPTION_CONTAINER/shellcont/shell").currentCellRow = 5
    session.findById("wnd[2]/usr/cntlOPTION_CONTAINER/shellcont/shell").selectedRows = "5"
    session.findById("wnd[2]/usr/cntlOPTION_CONTAINER/shellcont/shell").doubleClickCurrentCell()
    session.findById("wnd[1]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN001-LOW").text = "PSAPUNDO"
    session.findById("wnd[1]/tbar[0]/btn[0]").press()

    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").setCurrentCell (-1,"TOT_PER_USED")
    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").selectColumn ("TOT_PER_USED")
    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").contextMenu()
    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").selectContextMenuItem ("&FILTER")
    session.findById("wnd[1]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN002-LOW").text = "91"
    session.findById("wnd[1]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN002-HIGH").text = "100"
    session.findById("wnd[1]/tbar[0]/btn[0]").press()

    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").currentCellColumn = "TOT_PER_USED"
    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").selectColumn ("TOT_PER_USED")
    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").contextMenu()
    session.findById("wnd[0]/usr/tabsTS_TABSTRIP/tabpTS_TABSTRIP_TAB1_/ssub221_SCA:SAPLS_ORA_COCKPIT_5:0221/cntlCC_TAB1_0221/shellcont/shell").selectContextMenuItem ("&FIND")
    session.findById("wnd[1]/usr/chkGS_SEARCH-EXACT_WORD").selected = True
    session.findById("wnd[1]/usr/chkGS_SEARCH-SHOW_HITS").selected = True
    session.findById("wnd[1]/usr/txtGS_SEARCH-VALUE").text = "*"
    session.findById("wnd[1]/tbar[0]/btn[0]").press()
    session.findById("wnd[1]/usr/txtGS_SEARCH-SEARCH_INFO").setFocus()
    tablespace_details = session.findById("wnd[1]/usr/txtGS_SEARCH-SEARCH_INFO").text
    try : 
        numbers = re.findall(r'\d+', tablespace_details)
        tablespace_greater_than_90 = numbers[-1] 
    except :
        tablespace_greater_than_90 = 0
    #print("Tablespace > 90 :", tablespace_greater_than_90)
    session.findById("wnd[1]").close()

    session.findById("wnd[0]/tbar[0]/okcd").text = "/nex"
    session.findById("wnd[0]").sendVKey (0)

    from openpyxl import load_workbook
    Excel_path = os.path.abspath(os.path.join(current_directory, 'Scripts_run','Manual_Monitoring','Reports','Manual_Monitoring_Report.xlsx'))
    wb = load_workbook(Excel_path)
    ws = wb.active
    values = [int(sm12_previous_locks), int(sm12_today_locks), int(sm12_total_locks), int(st22_today_runtime_errors), int(sm37_total_job_time), sm51_status, int(db02_percentage), int(tablespace_greater_than_90)]

    # Write values into cells B4 to I4
    for index, value in enumerate(values):
        ws.cell(row=4, column=index+2, value=value) 

    wb.save(Excel_path)
    print(f"Excel Saved at {Excel_path}")
#'''
if __name__ == "__main__":
    run_pk9_script()