import pandas as pd
from tqdm import tqdm
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side
from math import ceil
import warnings
import sys
import os

current_directory = os.getcwd()
parameter_lists_directory = os.path.abspath(os.path.join(current_directory, 'Files', 'Common','Perimeter_Lists'))
patching_directory = os.path.abspath(os.path.join(current_directory, 'Scripts_run', 'Monday_Patching'))

file1_path = os.path.abspath(os.path.join(parameter_lists_directory, 'SAP Global Perimeter V18.4.xlsx'))
file2_path = os.path.abspath(os.path.join(parameter_lists_directory, 'MII Landscape details - Global_2024_06_10.xlsx'))

change_task_path = os.path.abspath(os.path.join(current_directory, 'Scripts_run','Thursday_Patching','Input','change_task.xlsx'))
servers_and_crq_path = os.path.abspath(os.path.join(current_directory, 'Scripts_run','Thursday_Patching','Input','Servers_&_CRQ_names.xlsx'))

warnings.filterwarnings("ignore", category=UserWarning, message="Workbook contains no default style*")

def search_name_in_sheet(workbook_path, sheet_name, name):
    df = pd.read_excel(workbook_path, sheet_name=sheet_name)
    if df.apply(lambda row: row.astype(str).str.lower().str.contains(name.lower(), case=False).any(), axis=1).any():
        return True
    return False

def search_name_in_workbook(workbook_path, name):
    workbook = pd.ExcelFile(workbook_path)
    for sheet_name in workbook.sheet_names:
        if search_name_in_sheet(workbook_path, sheet_name, name):
            return True
    return False

def extract_stop_start_details(change, description, file_path):
    df = pd.read_excel(file_path)
    match = df[(df['Change request'] == change) & (df['Short description'].str.contains(description))]
    if not match.empty:
        return {
            "Ctask": match.iloc[0]['Number'],
            "Ctask Assignee": "",
            "Planned start date": match.iloc[0]['Planned start date'],
            "Planned end date": match.iloc[0]['Planned end date']
        }
    else:
        return None

def extract_details(server_name, file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    name_row = df[df.apply(lambda row: row.astype(str).str.lower().str.contains(server_name.lower(), case=False).any(), axis=1)]

    if not name_row.empty:
        if file_path == file1_path and sheet_name == 'Inventory':
            sid = name_row.iloc[0, 1] if pd.isna(name_row.iloc[0, 2]) else name_row.iloc[0, 2]
            environment = name_row.iloc[0, 10]
            server_type = name_row.iloc[0, 12]
            dns_alias = name_row.iloc[0,7]

        elif file_path == file2_path and sheet_name == 'Local Sites':
            sid = name_row.iloc[0, 4]
            environment = ''
            server_type = ''
            dns_alias = ''
        elif file_path == file2_path and sheet_name == 'AWS':
            sid = name_row.iloc[0, 13]
            environment = ''
            server_type = name_row.iloc[0, 3]
            dns_alias = ''
        else:
            sid = ''
            environment = ''
            server_type = ''
            dns_alias = ''
        
        return sid, environment, server_type, dns_alias

    return '', '', '', ''

df = pd.read_excel(servers_and_crq_path)

current_change = None
current_servers = []

result = []

with tqdm(total=len(df), desc="Processing Servers", unit="row", colour='green') as pbar:
    for index, row in df.iterrows():
        server = str(row['Servers']).strip()
        change = row['CRQ']

        if pd.notna(change):
            if current_change:
                change_exists = any(current_change in item for item in result)
                if change_exists:
                    for item in result:
                        if current_change in item:
                            item[current_change].extend(current_servers)
                            break
                else:
                    change_servers = {current_change: current_servers}
                    result.append(change_servers)

            current_change = change
            current_servers = []

        if pd.notna(server):
            current_servers.append(server)

        pbar.update(1)

if current_change:
    change_exists = any(current_change in item for item in result)
    if change_exists:
        for item in result:
            if current_change in item:
                item[current_change].extend(current_servers)
                break
    else:
        change_servers = {current_change: current_servers}
        result.append(change_servers)

need_to_patch = []
closed = []

with tqdm(total=len(result), desc="Searching Servers", unit="change", colour='blue') as pbar:
    for change_servers in result:
        change = list(change_servers.keys())[0]
        servers = change_servers[change]

        with tqdm(total=len(servers), desc=f"Searching servers for {change}", unit="server", colour='blue') as inner_pbar:
            servers_found = []

            for server in servers:
                found_in_file1 = search_name_in_workbook(file1_path, server)
                if found_in_file1:
                    servers_found.append(server)
                else:
                    found_in_file2 = search_name_in_workbook(file2_path, server)
                    if found_in_file2:
                        servers_found.append(server)
                inner_pbar.update(1)

            if servers_found:
                need_to_patch.append({change: servers_found})
            else:
                closed.append(change)


output_path = os.path.abspath(os.path.join(current_directory, 'Scripts_run','Thursday_Patching','Output'))
if not os.path.exists(output_path):
        os.makedirs(output_path)

output_file_path = os.path.abspath(os.path.join(output_path,'output.txt'))

with open(output_file_path, 'w') as file:
    file.write("Close\n-----\n")
    for item in closed:
        file.write(item + "\n")

    file.write("\nNeed to Patch\n-------------\n")
    for item in need_to_patch:
        change, servers = list(item.items())[0]
        if servers:
            file.write("\n".join(servers) + "  " + change + "\n\n")
        else:
            file.write(change + "\n\n")

print(f"Output saved to: {output_file_path}")

data = []

with tqdm(total=len(need_to_patch), desc="Generating Excel", unit="change", colour='green') as pbar:
    for item in need_to_patch:
        for change, servers in item.items():
            if len(servers) == 1:
                servers.append('')  

            for server in servers:
                if server:  
                    if server == servers[-1]: 
                        short_description = "START"
                        description_type = "Start SAP Application"
                    else:
                        short_description = "STOP"
                        description_type = "Stop SAP"

                    sid, environment, server_type, dns_alias = extract_details(server, file1_path, 'Inventory')
                    if not sid:
                        sid, environment, server_type, dns_alias = extract_details(server, file2_path, 'Local Sites')
                    if not sid:
                        sid, environment, server_type, dns_alias = extract_details(server, file2_path, 'AWS')

                    # Extract details based on short description type
                    if short_description == "START":
                        details = extract_stop_start_details(change, description_type, change_task_path)
                    else:
                        details = extract_stop_start_details(change, description_type, change_task_path)

                    if details:
                        data.append({
                            "Change": change,
                            "Ctask": details["Ctask"],
                            "Ctask Assignee": details["Ctask Assignee"],
                            "Planned start date": details["Planned start date"],
                            "Planned end date": details["Planned end date"],
                            "Short description": short_description,
                            "SID": sid,
                            "Purpose": environment,
                            "Hostnames": server,
                            "DNS Alias": dns_alias
                        })
                    else:
                        data.append({
                            "Change": change,
                            "Ctask": "",
                            "Ctask Assignee": "",
                            "Planned start date": "",
                            "Planned end date": "",
                            "Short description": short_description,
                            "SID": sid,
                            "Purpose": environment,
                            "Hostnames": server,
                            "DNS Alias": dns_alias
                        })
                else:
                    data.append({
                        "Change": change,
                        "Ctask": "",
                        "Ctask Assignee": "",
                        "Planned start date": "",
                        "Planned end date": "",
                        "Short description": "",
                        "SID": '',
                        "Purpose": '',
                        "Hostnames": '',
                        "DNS Alias": ''
                    })
        pbar.update(1)

output_path = os.path.abspath(os.path.join(output_path,'need_to_patch.xlsx'))

with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
    df = pd.DataFrame(data)
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    workbook = writer.book
    sheet = workbook.active

    # Resize columns
    column_widths = {
        'A': 15,  # Change
        'B': 12,  # Ctask
        'C': 20,  # Ctask Assignee
        'D': 18,  # Planned start date
        'E': 18,  # Planned end date
        'F': 12,  # Short description
        'G': 12,  # SID
        'H': 15,  # Purpose
        'I': 20,  # Hostnames
        'J': 20   # DNS Alias
    }

    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width
        for row in range(1, len(df) + 2):  
            sheet.cell(row=row, column=sheet[col + '1'].column).border = Border(left=Side(style='thin'), right=Side(style='thin'))


    current_change = None
    start_row = 2

    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        change_value = row[0]
        short_desc = row[5]
        if change_value != current_change:
            if current_change is not None:
                sheet.merge_cells(start_row=start_row, start_column=1, end_row=i-1, end_column=1)
                cell = sheet.cell(row=start_row, column=1)
                cell.alignment = Alignment(vertical='center')

                for row_num in range(start_row, i): 
                    if row_num == start_row and len(df) > 1:  
                        sheet.cell(row=row_num, column=6).value = "STOP"
                        
                        stop_data = extract_stop_start_details(current_change, "Stop SAP", change_task_path)
                        if stop_data:
                            sheet.cell(row=row_num, column=2).value = stop_data["Ctask"]
                            sheet.cell(row=row_num, column=3).value = stop_data["Ctask Assignee"]
                            sheet.cell(row=row_num, column=4).value = stop_data["Planned start date"]
                            sheet.cell(row=row_num, column=5).value = stop_data["Planned end date"]
                    else:
                        sheet.cell(row=row_num, column=6).value = "START"
                        
                        start_data = extract_stop_start_details(current_change, "Start SAP Application", change_task_path)
                        if start_data:
                            sheet.cell(row=row_num, column=2).value = start_data["Ctask"]
                            sheet.cell(row=row_num, column=3).value = start_data["Ctask Assignee"]
                            sheet.cell(row=row_num, column=4).value = start_data["Planned start date"]
                            sheet.cell(row=row_num, column=5).value = start_data["Planned end date"]

               
                for col_num in range(2, 7): 
                    col = get_column_letter(col_num)
                    total_rows = i - start_row
                    half_rows = ceil(total_rows / 2)
                    sheet.merge_cells(start_row=start_row, start_column=col_num, end_row=start_row + half_rows - 1, end_column=col_num)
                    sheet.merge_cells(start_row=start_row + half_rows, start_column=col_num, end_row=i - 1, end_column=col_num)
                    for r in range(start_row, i):
                        sheet.cell(row=r, column=col_num).alignment = Alignment(vertical='center')

                    start_cell = sheet.cell(row=start_row, column=col_num)
                    end_cell_group1 = sheet.cell(row=start_row + half_rows - 1, column=col_num)
                    end_cell_group2 = sheet.cell(row=i - 1, column=col_num)
                    start_cell.border = Border(top=Side(style='thick', color='00FF00'), bottom=Side(style='thin'), right=Side(style='thin'))
                    end_cell_group1.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
                    end_cell_group2.border = Border(top=Side(style='thin'), bottom=Side(style='thick', color='00FF00'), right=Side(style='thin'))

            current_change = change_value
            start_row = i

    
    if current_change is not None:
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=len(df) + 1, end_column=1)
        cell = sheet.cell(row=start_row, column=1)
        cell.alignment = Alignment(vertical='center')

        
        for row_num in range(start_row, len(df) + 1):  
            if row_num == start_row and len(df) > 1:  
                sheet.cell(row=row_num, column=6).value = "STOP"
                stop_data = extract_stop_start_details(current_change, "Stop SAP", change_task_path)
                if stop_data:
                    sheet.cell(row=row_num, column=2).value = stop_data["Ctask"]
                    sheet.cell(row=row_num, column=3).value = stop_data["Ctask Assignee"]
                    sheet.cell(row=row_num, column=4).value = stop_data["Planned start date"]
                    sheet.cell(row=row_num, column=5).value = stop_data["Planned end date"]
            else:
                sheet.cell(row=row_num, column=6).value = "START"
                start_data = extract_stop_start_details(current_change, "Start SAP Application", change_task_path)
                if start_data:
                    sheet.cell(row=row_num, column=2).value = start_data["Ctask"]
                    sheet.cell(row=row_num, column=3).value = start_data["Ctask Assignee"]
                    sheet.cell(row=row_num, column=4).value = start_data["Planned start date"]
                    sheet.cell(row=row_num, column=5).value = start_data["Planned end date"]

        
        for col_num in range(2, 7): 
            col = get_column_letter(col_num)
            total_rows = len(df) - start_row + 1
            half_rows = ceil(total_rows / 2)
            end_row_group1 = start_row + half_rows - 1
            start_row_group2 = start_row + half_rows
            end_row_group2 = len(df) + 1
            sheet.merge_cells(start_row=start_row, start_column=col_num, end_row=end_row_group1, end_column=col_num)
            sheet.merge_cells(start_row=start_row_group2, start_column=col_num, end_row=end_row_group2, end_column=col_num)
            for r in range(start_row, end_row_group1 + 1):
                sheet.cell(row=r, column=col_num).alignment = Alignment(vertical='center')
            for r in range(start_row_group2, end_row_group2 + 1):
                sheet.cell(row=r, column=col_num).alignment = Alignment(vertical='center')

            start_cell = sheet.cell(row=start_row, column=col_num)
            end_cell_group1 = sheet.cell(row=end_row_group1, column=col_num)
            end_cell_group2 = sheet.cell(row=end_row_group2, column=col_num)

            start_cell.border = Border(top=Side(style='thick', color='00FF00'), bottom=Side(style='thin'), right=Side(style='thin'))
            end_cell_group1.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'))
            end_cell_group2.border = Border(top=Side(style='thin'), bottom=Side(style='thick', color='00FF00'), right=Side(style='thin'))


    current_change = None
    start_row = 2
    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        change_value = row[0]
        if change_value != current_change:
            if current_change is not None:
                
                sheet.cell(row=i-1, column=1).border = Border(bottom=Side(style='thick', color='00FF00'), right=Side(style='thin'))
                
                for row_num in range(start_row, len(df) + 1):
                    for col_num in range(7, 10):
                        sheet.cell(row=row_num, column=col_num).border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
                    sheet.cell(row=row_num, column=9).border = Border(bottom=Side(style='thin'), right=Side(style='thick', color='00FF00'))
                
                for col_num in range(7, 10): 
                    sheet.cell(row=i-1, column=col_num).border = Border(bottom=Side(style='thick', color='00FF00'), right=Side(style='thin'))
                sheet.cell(row=i-1, column=9).border = Border(bottom=Side(style='thick', color='00FF00'), right=Side(style='thick', color='00FF00'))
            current_change = change_value
            start_row = i

    
    if current_change is not None:
        
        sheet.cell(row=len(df) + 1, column=1).border = Border(bottom=Side(style='thick', color='00FF00'), right=Side(style='thin'))
        
        for row_num in range(start_row, len(df) + 1):
            for col_num in range(7, 10):
                sheet.cell(row=row_num, column=col_num).border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
            sheet.cell(row=row_num, column=9).border = Border(bottom=Side(style='thin'), right=Side(style='thick', color='00FF00'))
        for col_num in range(7, 10): 
            sheet.cell(row=len(df) + 1, column=col_num).border = Border(bottom=Side(style='thick', color='00FF00'), right=Side(style='thin'))
        sheet.cell(row=len(df) + 1, column=9).border = Border(bottom=Side(style='thick', color='00FF00'), right=Side(style='thick', color='00FF00'))


    
        
print(f"Excel file created: {output_path}")
