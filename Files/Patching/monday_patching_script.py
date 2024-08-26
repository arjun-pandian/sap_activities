import pandas as pd
from tqdm import tqdm
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import os

current_directory = os.getcwd()
parameter_lists_directory = os.path.abspath(os.path.join(current_directory, 'Files', 'Common','Perimeter_Lists'))
patching_directory = os.path.abspath(os.path.join(current_directory, 'Scripts_run', 'Monday_Patching'))

ao_names_directory = os.path.abspath(os.path.join(current_directory, 'Scripts_run','Monday_Patching','Output','AO_Names.xlsx'))

output_patching_directory = os.path.abspath(os.path.join(current_directory, 'Scripts_run','Monday_Patching','Output'))
if not os.path.exists(output_patching_directory):
        os.makedirs(output_patching_directory)

def create_ao_name_excel(file_path):
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet1'

    columns = ["Servers", "Sanofi application AO Name"]
    sheet.append(columns)

    fill = PatternFill(start_color='006400', end_color='006400', fill_type='solid')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color='FFFFFF')
        cell.border = border
    
    column_widths = [10, 30]
    
    for i, width in enumerate(column_widths, start=1):
        column_letter = get_column_letter(i)
        sheet.column_dimensions[column_letter].width = width

    workbook.save(file_path)

    print(f"Excel file '{file_path}' has been created/replaced with the specified format.")


create_ao_name_excel(ao_names_directory)


def search_name_in_sheet(workbook_path, sheet_name, name):
    df = pd.read_excel(workbook_path, sheet_name=sheet_name)
    if df.apply(lambda row: row.astype(str).str.lower().str.contains(name.lower(), case=False).any(), axis=1).any():
        return name
    return None

def search_name_in_workbook(workbook_path, name):
    workbook = pd.ExcelFile(workbook_path)
    for sheet_name in workbook.sheet_names:
        found_name = search_name_in_sheet(workbook_path, sheet_name, name)
        if found_name:
            return found_name, workbook_path, sheet_name
    return None, None, None

file1_path = os.path.abspath(os.path.join(parameter_lists_directory, 'SAP Global Perimeter V18.4.xlsx'))
file2_path = os.path.abspath(os.path.join(parameter_lists_directory, 'MII Landscape details - Global_2024_06_10.xlsx'))

server_names_path = os.path.abspath(os.path.join(patching_directory, 'Input','server_names.txt'))

with open(server_names_path, 'r') as file:
    names_to_search = [line.strip().lower() for line in file]
    total_names = len(names_to_search)
    print("Total names to search:", total_names)

names_found = []
duplicate_names = defaultdict(int)
with tqdm(total=total_names, desc="Searching", unit="names") as pbar:
    for name in names_to_search:
        found_in_file1, file_path, sheet_name = search_name_in_workbook(file1_path, name)
        if found_in_file1:
            names_found.append((found_in_file1, file_path, sheet_name))
            duplicate_names[found_in_file1] += 1
        else:
            found_in_file2, file_path, sheet_name = search_name_in_workbook(file2_path, name)
            if found_in_file2:
                names_found.append((found_in_file2, file_path, sheet_name))
                duplicate_names[found_in_file2] += 1
        pbar.update(1)

print("\nTotal names found:", len(names_found))

total_duplicates = sum(1 for count in duplicate_names.values() if count > 1)
print(f"Total duplicates found: {total_duplicates}")

extracted_data = []
first_name_found = False

for name, file_path, sheet_name in names_found:
    if file_path == file1_path and sheet_name == 'Inventory':
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        name_row = df[df.apply(lambda row: row.astype(str).str.lower().str.contains(name.lower(), case=False).any(), axis=1)]
        if not name_row.empty:
            sid = name_row.iloc[0, 2]
            environment = name_row.iloc[0, 4]
            server_type = name_row.iloc[0, 12]
            extracted_data.append({
                'Servers': name,
                'SID': sid,
                'Environment': environment,
                'Type': server_type,
                'Sanofi application AO Name': ''
            })
            first_name_found = True
    else:
        extracted_data.append({
            'Servers': name,
            'SID': '',
            'Environment': '',
            'Type': '',
            'Sanofi application AO Name': ''
        })

if not first_name_found:
    extracted_data = [data for data in extracted_data if data['Servers'] != '']

extracted_df = pd.DataFrame(extracted_data)

output_excel_path = os.path.abspath(os.path.join(output_patching_directory, 'servers_output.xlsx'))
extracted_df.to_excel(output_excel_path, index=False)

wb = load_workbook(output_excel_path)
ws = wb.active

column_widths = {
    'A': 15,
    'B': 8,
    'C': 15,
    'D': 10,
    'E': 25
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

header_fill = PatternFill(start_color='71AD70', end_color='71AD70', fill_type='solid')
header_font = Font(bold=True, color='000000')
border = Border(left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000'))
for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
for row_idx, data in enumerate(extracted_data, start=2):
    cell = ws.cell(row=row_idx, column=5)
    cell.value = f'=VLOOKUP(A{row_idx},[AO_Names.xlsx]Sheet1!A:B, 2, FALSE)'

    cell.border = border

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = border
wb.save(output_excel_path)

print(f"\nExtracted data saved to '{output_excel_path}' successfully.")


summary_file_path = os.path.abspath(os.path.join(output_patching_directory,'summary.txt'))
with open(summary_file_path, 'w') as summary_file:
    summary_file.write(f"Total number of servers: {total_names}\n")
    summary_file.write(f"Total names found: {len(names_found)}\n\n")
    summary_file.write("Names found and their locations:\n")
    for name, file_path, sheet_name in names_found:
        summary_file.write(f"{name} found in {file_path}, sheet: {sheet_name}\n")
    summary_file.write(f"\nTotal duplicates found: {total_duplicates}\n\n")
    if (total_duplicates != 0) :
        summary_file.write("Duplicate names and their occurrences:\n")
        for name, count in duplicate_names.items():
            if count > 1:
                summary_file.write(f"{name}: {count} occurrences\n")

print(f"Summary information saved to '{summary_file_path}' successfully.")
